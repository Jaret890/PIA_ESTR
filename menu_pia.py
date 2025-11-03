import sqlite3
import datetime
import os
import csv
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment

BD_NOMBRE = "coworking.db"
FORMATO_FECHA = "%m/%d/%Y"

def fecha_a_timestamp(fecha_str):
    fecha =  datetime.datetime.strptime(fecha_str, FORMATO_FECHA)
    return fecha.timestamp()

def timestamp_a_fecha(ts):
    return datetime.datetime.fromtimestamp(ts).strftime(FORMATO_FECHA)

def fecha_valida(fecha_str):
    try:
        fecha = datetime.datetime.strptime(fecha_str, FORMATO_FECHA).date()
        hoy = datetime.date.today()
        delta = (fecha - hoy).days
        if delta < 2:
            print("--La fecha debe ser al menos dos dias posterior a hoy--")
            return False
        if fecha.weekday() == 6:
            lunes = fecha + datetime.timedelta(days=1)
            print(f"--La fecha cae en domingo. Se propone el lunes siguiente ({lunes.strftime(FORMATO_FECHA)})--")
            return False
        return True
    except:
        print(f"--Formato invalido usa el formato {FORMATO_FECHA} (ejemplo: {datetime.date.today().strftime(FORMATO_FECHA)})")
        return False
    
def turnos(turno):
    return turno.lower() in ["mañana", "tarde", "noche"]

def registrar_cliente():
    nombre = input("Nombre del cliente: ")
    apellidos = input("Apellidos: ")
    if not nombre or not apellidos:
        print("--Nombre y apellidos son obligatorios--")
        return
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute("INSERT INTO clientes (nombre, apellidos) VALUES (?, ?)", (nombre, apellidos))
    conn.commit()
    cid = cursor.lastrowid
    conn.close()
    print(f"--Cliente registrado con ID: {cid}--")


def listar_clientes():
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute("SELECT id_cliente, nombre, apellidos FROM clientes ORDER BY apellidos, nombre")
    clientes = cursor.fetchall()
    conn.close()
    print("--Clientes registrados:")
    for c in clientes:
        print(f"{c[0]} - {c[2]}, {c[1]}")
    return [c[0] for c in clientes]

def registrar_sala():
    sala = input("--Nombre de la sala: ")
    try:
        cupo = int(input("--Cupo de la sala: "))
        if cupo <= 0:
            raise ValueError
    except:
        print("--El cupo debe de ser positivo--")
        return
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute("INSERT INTO salas (nombre, cupo) VALUES (?, ?)", (sala, cupo))
    conn.commit()
    sid = cursor.lastrowid
    conn.close()
    print(f"--Sala registrada con ID: {sid}--")

def listar_salas():
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute("SELECT id_sala, nombre, cupo FROM salas ORDER BY nombre")
    salas = cursor.fetchall()
    conn.close()
    print("--Salas registradas:")
    for s in salas:
        print(f"{s[0]} - {s[1]} (Cupo: {s[2]})")
    return [s[0] for s in salas]

def registrar_reservacion():
    claves_clientes = listar_clientes()
    if not claves_clientes:
        print("--No hay clientes registrados--")
        return

    try:
        id_cliente = int(input("--ID del cliente: "))
    except:
        print("--ID invalido--")
        return
    if id_cliente not in claves_clientes:
        print("--Cliente no encontrado--")
        return
    
    fecha = input(f"Fecha ({FORMATO_FECHA}): ")
    if not fecha_valida(fecha):
        return
    
    ts = fecha_a_timestamp(fecha)
    claves_salas = listar_salas()

    if not claves_salas:
        print("--No hay salas registradas--")
        return
    
    try:
        id_sala = int(input("--ID de la sala: "))
    except:
        print("--ID invalido--")
        return
    if id_sala not in claves_salas:
        print("--Sala no valida--")
        return
    
    turno = input("--Turno (mañana/tarde/noche): ").strip().lower()
    if not turnos(turno):
        print("--Turno no valido--")
        return
    
    evento = input("--Nombre del evento: ")
    if not evento or evento.isspace():
        print("--Nombre del evento no valido--")
        return
    
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute("SELECT 1 FROM reservaciones WHERE id_sala=? AND fecha=? AND turno=?", (id_sala, ts, turno))
    if cursor.fetchone():
        print("--Ya hay una reservación en ese turno para esa sala--")
        conn.close()
        return

    cursor.execute("INSERT INTO reservaciones (id_cliente, id_sala, fecha, turno, evento) VALUES (?, ?, ?, ?, ?, 'activa')",
                   (id_cliente, id_sala, ts, turno, evento))
    conn.commit()
    rid = cursor.lastrowid
    conn.close()
    print(f"--Reservación registrada con folio: {rid}--")

def consultar_por_fecha():
    fecha = input(f"Fecha a consultar ({FORMATO_FECHA}) o ENTER para hoy: ").strip()
    if not fecha:
        fecha = datetime.date.today().strftime(FORMATO_FECHA)
    ts = fecha_a_timestamp(fecha)

    conn = conectar()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT r.id_reservacion, r.fecha, r.turno, r.evento,
               c.nombre || ' ' || c.apellidos AS cliente,
               s.nombre AS sala, s.cupo
        FROM reservaciones r
        JOIN clientes c ON r.id_cliente = c.id_cliente
        JOIN salas s ON r.id_sala = s.id_sala
        WHERE r.fecha = ? AND r.estado = 'activa'
        ORDER BY r.turno, s.nombre
    ''', (ts,))
    reservas = cursor.fetchall()
    conn.close()

    if not reservas:
        print("--No hay reservaciones para esa fecha--")
        return

    print(f"--Reservaciones del {fecha}:")
    print("-" * 70)
    for r in reservas:
        print(f"Folio: {r[0]} | Fecha: {timestamp_a_fecha(r[1])} | Turno: {r[2]} | Evento: {r[3]} | Cliente: {r[4]} | Sala: {r[5]} (Cupo {r[6]})")
    print("-" * 70)

    exportar_opcion(reservas, fecha)

def exportar_opcion(reservas, fecha):
    print("--¿Desea exportar el reporte?--")
    print("1.CSV")
    print("2.JSON")
    print("3.Excel")
    print("4.No exportar")

    opcion = input("--Opcion: ")
    
    if opcion == "1":
        exportar_csv(reservas, f"reporte_{fecha.replace('/','-')}.csv")
    elif opcion == "2":
        exportar_json(reservas, f"reporte_{fecha.replace('/','-')}.json")
    elif opcion == "3":
        exportar_excel(reservas, f"reporte_{fecha.replace('/','-')}.xlsx")
    else:
        print("--No se exporto el reporte--")

def exportar_csv(reservas,nombre_archivo):
    with open(nombre_archivo, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Folio", "Fecha", "Turno", "Evento", "Cliente", "Sala", "Cupo"])
        for r in reservas:
            writer.writerow([r[0], timestamp_a_fecha(r[1]), r[2], r[3], r[4], r[5], r[6]])
    print(f"--Reporte exportado a {nombre_archivo}--")

def exportar_json(reservas, nombre_archivo):
    data = [
        {"Folio": r[0], "Fecha": timestamp_a_fecha(r[1]), "Turno": r[2], "Evento": r[3],
         "Cliente": r[4], "Sala": r[5], "Cupo": r[6]} for r in reservas
    ]
    with open(nombre_archivo, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=4)
    print(f"--Reporte exportado a {nombre_archivo}--")

def exportar_excel(reservas, nombre_archivo):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reservaciones"
    ws["A1"] = "Reporte de reservaciones"
    ws["A1"].font = Font(bold=True, size=14)

    headers = ["Folio", "Fecha", "Turno", "Evento", "Cliente", "Sala", "Cupo"]
    ws.append([])
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(bottom=Side(style="thick"))

    for r in reservas:
        ws.append([r[0], timestamp_a_fecha(r[1]), r[2], r[3], r[4], r[5], r[6]])

    for row in ws.iter_rows(min_row=4, max_row=3 + len(reservas), min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(horizontal="center")

    wb.save(nombre_archivo)
    print(f"--Reporte exportado a {nombre_archivo}--")

def listar_reservaciones():
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute("SELECT id_reservacion, evento FROM reservaciones ORDER BY id_reservacion")
    datos = cursor.fetchall()
    conn.close()
    if not datos:
        print("--No hay reservaciones registradas--")
        return []
    print("--Reservaciones disponibles:")
    for r in datos:
        print(f"Folio: {r[0]} - Evento: {r[1]}")
    return [r[0] for r in datos]

def editar_evento():
    folios = listar_reservaciones()
    if not folios:
        return
    try:
        folio = int(input("Folio del evento a editar: ").strip())
    except:
        print("Folio inválido.")
        return
    if folio not in folios:
        print("Folio no existe.")
        return

    conn = conectar()
    cursor = conn.cursor()
    cursor.execute("SELECT evento FROM reservaciones WHERE id_reservacion=?", (folio,))
    data = cursor.fetchone()
    print(f"Evento actual: {data[0]}")
    nuevo = input("Nuevo nombre del evento: ").strip()
    if not nuevo:
        print("--El nombre no puede estar vacío--")
        conn.close()
        return
    cursor.execute("UPDATE reservaciones SET evento=? WHERE id_reservacion=?", (nuevo, folio))
    conn.commit()
    conn.close()
    print("--Evento actualizado correctamente--")

def cancelar_reservacion():
    print("--Cancelar una reservación--")
    fecha_inicio = input(f"Fecha inicial ({FORMATO_FECHA}): ")
    fecha_fin = input(f"Fecha final ({FORMATO_FECHA}): ")

    try:
        ts_inicio = fecha_a_timestamp(fecha_inicio)
        ts_fin = fecha_a_timestamp(fecha_fin)
    except:
        print("--Formato de fecha incorrecto--")
        return

    conn = conectar()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT r.id_reservacion, r.fecha, r.evento, r.turno,
               c.nombre || ' ' || c.apellidos AS cliente,
               s.nombre AS sala
        FROM reservaciones r
        JOIN clientes c ON r.id_cliente = c.id_cliente
        JOIN salas s ON r.id_sala = s.id_sala
        WHERE r.fecha BETWEEN ? AND ? AND r.estado='activa'
        ORDER BY r.fecha
    ''', (ts_inicio, ts_fin))
    reservas = cursor.fetchall()

    if not reservas:
        print("--No hay reservaciones activas en ese rango de fechas--")
        conn.close()
        return

    print("--Reservaciones encontradas:")
    for r in reservas:
        print(f"Folio: {r[0]} | Fecha: {timestamp_a_fecha(r[1])} | Evento: {r[2]} | Turno: {r[3]} | Cliente: {r[4]} | Sala: {r[5]}")

    try:
        folio = int(input("--Ingresa el folio de la reservación a cancelar: "))
    except:
        print("--Folio inválido--")
        conn.close()
        return

    cursor.execute("SELECT fecha FROM reservaciones WHERE id_reservacion=? AND estado='activa'", (folio,))
    data = cursor.fetchone()
    if not data:
        print("--No se encontró una reservación activa con ese folio--")
        conn.close()
        return

    fecha_evento = datetime.datetime.fromtimestamp(data[0]).date()
    hoy = datetime.date.today()
    delta = (fecha_evento - hoy).days
    if delta < 2:
        print("--Solo se pueden cancelar reservaciones con al menos dos días de anticipación--")
        conn.close()
        return

    confirmar = input("--¿Seguro que deseas cancelar esta reservación? (s/n): ").upper()
    if confirmar != "S":
        print("--Operación cancelada--")
        conn.close()
        return

    cursor.execute("UPDATE reservaciones SET estado='cancelada' WHERE id_reservacion=?", (folio,))
    conn.commit()
    conn.close()
    print("--Reservación cancelada correctamente--")

def conectar():
    return sqlite3.connect(BD_NOMBRE)

def inicializar_bd():
    existe = os.path.exists(BD_NOMBRE)
    conn = conectar()
    cursor = conn.cursor()

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS clientes (
            id_cliente INTEGER PRIMARY KEY,
            nombre TEXT NOT NULL,
            apellidos TEXT NOT NULL
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS salas (
            id_sala INTEGER PRIMARY KEY,
            nombre TEXT NOT NULL,
            cupo INTEGER NOT NULL CHECK(cupo > 0)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS reservaciones (
            id_reservacion INTEGER PRIMARY KEY,
            id_cliente INTEGER NOT NULL,
            id_sala INTEGER NOT NULL,
            fecha REAL NOT NULL,  -- timestamp
            turno TEXT NOT NULL,
            evento TEXT NOT NULL,
            estado TEXT DEFAULT 'activa',
            FOREIGN KEY(id_cliente) REFERENCES clientes(id_cliente),
            FOREIGN KEY(id_sala) REFERENCES salas(id_sala),
            UNIQUE(id_sala, fecha, turno)
        )
    ''')

    conn.commit()
    conn.close()

    if existe:
        print("--Se cargó el estado anterior de la base de datos--")
    else:
        print("--No se encontró una versión anterior. Se inicia con un estado vacío--")

def menu():
    inicializar_bd()
    print("--Sistema de Reservas de Coworking--")
    while True:
        print("--Menu--")
        print("1.- Registrar reservacion")
        print("2.- Editar nombre del evento")
        print("3.- Consultar reservaciones por fecha")
        print("4.- Cancelar reservacion")
        print("5.- Registrar cliente")
        print("6.- Registrar sala")  
        print("7.- Salir")

        opcion = input("--Movimiento a realizar: ")
        if opcion == "1":
            registrar_reservacion()
        elif opcion == "2":
            editar_evento()    
        elif opcion == "3":
            consultar_por_fecha()
        elif opcion == "4":
            cancelar_reservacion()   
        elif opcion == "5":
            registrar_cliente()
        elif opcion == "6":
            registrar_sala()
        elif opcion == "7":
            confirmar = input("--¿Seguro que desea salir? (s/n): ").upper()
            if confirmar == "S":
                print("--Adios--")
                break
        else:
            print("--Opcion invalida--")
    
if __name__ == "__main__":
    menu()